import io
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─── Colores de marca ───────────────────────────────────────────────
GOLD = colors.HexColor('#C9A96E')
NUDE = colors.HexColor('#F5ECD7')
WHITE = colors.white
DARK = colors.HexColor('#2C2C2C')


def generate_finance_pdf(data: dict, report_type: str = 'monthly') -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    elements = []

    # Título
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontSize=22,
        textColor=DARK,
        alignment=TA_CENTER,
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=GOLD,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    label_style = ParagraphStyle(
        'Label',
        parent=styles['Normal'],
        fontSize=10,
        textColor=DARK,
        spaceAfter=6
    )

    elements.append(Paragraph('Seremyc Sthetic', title_style))
    elements.append(Paragraph('Reporte Financiero', subtitle_style))
    elements.append(Spacer(1, 10))

    if report_type == 'monthly':
        elements += _build_monthly_pdf(data, label_style, styles)
    else:
        elements += _build_annual_pdf(data, label_style, styles)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def _build_monthly_pdf(data, label_style, styles):
    elements = []

    # Resumen
    summary_data = [
        ['Concepto', 'Monto'],
        ['Total Ingresos', f"$ {data['incomes']:,.2f}"],
        ['Total Gastos', f"$ {data['expenses']:,.2f}"],
        ['Balance', f"$ {data['balance']:,.2f}"],
    ]

    summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GOLD),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, -2), NUDE),
        ('BACKGROUND', (0, -1), (-1, -1), GOLD),
        ('TEXTCOLOR', (0, -1), (-1, -1), WHITE),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [WHITE, NUDE]),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Detalle de registros
    if data.get('records'):
        elements.append(Paragraph('Detalle de Movimientos', label_style))
        elements.append(Spacer(1, 8))

        records_data = [['Fecha', 'Tipo', 'Categoría', 'Descripción', 'Monto']]
        for r in data['records']:
            records_data.append([
                r.get('date', ''),
                'Ingreso' if r.get('type') == 'income' else 'Gasto',
                r.get('category', ''),
                r.get('description', '')[:30] if r.get('description') else '',
                f"$ {r.get('amount', 0):,.2f}"
            ])

        records_table = Table(records_data, colWidths=[1.1*inch, 0.9*inch, 1.2*inch, 2*inch, 1.1*inch])
        records_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), GOLD),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, NUDE]),
            ('PADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(records_table)

    return elements


def _build_annual_pdf(data, label_style, styles):
    elements = []

    elements.append(Paragraph(f"Año {data['year']}", label_style))
    elements.append(Spacer(1, 10))

    # Resumen anual
    summary_data = [
        ['Total Ingresos', 'Total Gastos', 'Balance Anual'],
        [
            f"$ {data['total_income']:,.2f}",
            f"$ {data['total_expense']:,.2f}",
            f"$ {data['total_balance']:,.2f}"
        ]
    ]
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GOLD),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, -1), NUDE),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Tabla mensual
    elements.append(Paragraph('Resumen por Mes', label_style))
    elements.append(Spacer(1, 8))

    monthly_data = [['Mes', 'Ingresos', 'Gastos', 'Balance']]
    for m in data['months']:
        monthly_data.append([
            m['month_name'],
            f"$ {m['income']:,.2f}",
            f"$ {m['expense']:,.2f}",
            f"$ {m['balance']:,.2f}"
        ])

    monthly_table = Table(monthly_data, colWidths=[2*inch, 2*inch, 2*inch, 2*inch])
    monthly_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GOLD),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, NUDE]),
        ('PADDING', (0, 0), (-1, -1), 7),
    ]))
    elements.append(monthly_table)

    return elements


def generate_finance_excel(data: dict, report_type: str = 'monthly') -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active

    # Estilos
    gold_fill = PatternFill(start_color='C9A96E', end_color='C9A96E', fill_type='solid')
    nude_fill = PatternFill(start_color='F5ECD7', end_color='F5ECD7', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    bold_font = Font(bold=True, color='FFFFFF')
    dark_font = Font(bold=True, color='2C2C2C')
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    if report_type == 'monthly':
        ws.title = 'Reporte Mensual'
        _build_monthly_excel(ws, data, gold_fill, nude_fill, white_fill, bold_font, dark_font, center, thin_border)
    else:
        ws.title = 'Reporte Anual'
        _build_annual_excel(ws, data, gold_fill, nude_fill, white_fill, bold_font, dark_font, center, thin_border)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _build_monthly_excel(ws, data, gold_fill, nude_fill, white_fill, bold_font, dark_font, center, thin_border):
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = 'SEREMYC STHETIC - REPORTE FINANCIERO MENSUAL'
    ws['A1'].font = Font(bold=True, size=14, color='2C2C2C')
    ws['A1'].alignment = center

    # Resumen
    ws['A3'] = 'RESUMEN'
    ws['A3'].font = Font(bold=True, color='FFFFFF')
    ws['A3'].fill = gold_fill

    headers = ['Concepto', 'Monto']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = bold_font
        cell.fill = gold_fill
        cell.alignment = center
        cell.border = thin_border

    rows = [
        ['Total Ingresos', data['incomes']],
        ['Total Gastos', data['expenses']],
        ['Balance', data['balance']],
    ]
    for i, row in enumerate(rows, 5):
        fill = nude_fill if i % 2 == 0 else white_fill
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.alignment = center
            cell.border = thin_border

    # Detalle
    if data.get('records'):
        ws['A9'] = 'DETALLE DE MOVIMIENTOS'
        ws['A9'].font = Font(bold=True, color='FFFFFF')
        ws['A9'].fill = gold_fill

        detail_headers = ['Fecha', 'Tipo', 'Categoría', 'Descripción', 'Monto']
        for col, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=10, column=col, value=h)
            cell.font = bold_font
            cell.fill = gold_fill
            cell.alignment = center
            cell.border = thin_border

        for i, r in enumerate(data['records'], 11):
            fill = nude_fill if i % 2 == 0 else white_fill
            row_data = [
                r.get('date', ''),
                'Ingreso' if r.get('type') == 'income' else 'Gasto',
                r.get('category', ''),
                r.get('description', ''),
                r.get('amount', 0)
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.fill = fill
                cell.alignment = center
                cell.border = thin_border

    # Ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15


def _build_annual_excel(ws, data, gold_fill, nude_fill, white_fill, bold_font, dark_font, center, thin_border):
    # Título
    ws.merge_cells('A1:D1')
    ws['A1'] = f"SEREMYC STHETIC - REPORTE ANUAL {data['year']}"
    ws['A1'].font = Font(bold=True, size=14, color='2C2C2C')
    ws['A1'].alignment = center

    # Resumen anual
    summary_headers = ['Total Ingresos', 'Total Gastos', 'Balance Anual']
    for col, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = bold_font
        cell.fill = gold_fill
        cell.alignment = center
        cell.border = thin_border

    summary_values = [data['total_income'], data['total_expense'], data['total_balance']]
    for col, val in enumerate(summary_values, 1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.fill = nude_fill
        cell.alignment = center
        cell.border = thin_border
        cell.font = Font(bold=True, color='2C2C2C')

    # Tabla mensual
    ws['A6'] = 'RESUMEN MENSUAL'
    ws['A6'].font = bold_font
    ws['A6'].fill = gold_fill

    monthly_headers = ['Mes', 'Ingresos', 'Gastos', 'Balance']
    for col, h in enumerate(monthly_headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = bold_font
        cell.fill = gold_fill
        cell.alignment = center
        cell.border = thin_border

    for i, m in enumerate(data['months'], 8):
        fill = nude_fill if i % 2 == 0 else white_fill
        row_data = [m['month_name'], m['income'], m['expense'], m['balance']]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.alignment = center
            cell.border = thin_border

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18